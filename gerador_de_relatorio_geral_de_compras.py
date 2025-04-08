import pandas as pd
import math
import numpy as np
import time
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, Label, Button, Frame
from PIL import Image, ImageTk
import os
import sys



# Função para selecionar arquivos
def selecionar_arquivo(mensagem):
    # Cria uma janela root do tkinter (não será exibida)
    root = tk.Tk()
    root.withdraw()  # Esconde a janela principal do tkinter

    # Abre a janela de diálogo para selecionar o arquivo
    filename = filedialog.askopenfilename(
        title=mensagem,  # Título da janela
        filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]  # Tipos de arquivo permitidos
    )

    # Se o usuário cancelar a janela de diálogo, retorne None
    if not filename:
        print("\nSeleção de arquivo cancelada. Interrompendo a execução do código.")
        return None

    print(f'Arquivo selecionado: {filename}')
    return filename

# Função para se parar número de string
def estoque(valor):
    partes = valor.split()
    numero = partes[0].replace('.', '').replace(',', '.')
    numero_int = int(float(numero))
    unidade = partes[1] if len(partes) > 1 else ''
    return numero_int, unidade

# Função para se parar string de número
def custo(valor):
    return float(valor.replace('R$', '').replace('.', '').replace(',', '.').strip())

# Custo Adicional
custo_adicional = 1.14

# Fator de Reposição: representa o quanto quero ter no estoque em relação as vendas
fator = 1.1

# Função para processar o Catálogo de Produtos
def processar_catalogo_produtos(filename):
    print(f'\n1. Processando Catálogo de Produtos.')
    df = pd.read_excel(filename)
    df.columns = df.iloc[7]
    df = df.drop(range(0, 8)) # Linhas do topo desnecessárias
    df = df.iloc[:, [0, 1, 2, 4, 7, 8, 11]] # Seleciona as colunas desejadas
    df = df.rename(columns={df.columns[0]: "Cod BR", df.columns[1]: "Produto", df.columns[2]: "Cod PY", df.columns[5]: "Estoque BR", df.columns[6]: "Custo"}) # Corrige os nomes
    df = df.dropna(subset=['Produto']) # Exclui todas as linhas onde na coluna Produto estiver em branco 
    df[['Estoque BR', 'Unidade']] = df['Estoque BR'].apply(estoque).apply(pd.Series) # Separa número e texto
    df['Custo'] = df['Custo'].apply(custo) # Separa sifrão e valores
    df['Custo'] = df['Custo'] / custo_adicional # Para o custo ser igual ao preço do fornecedor
    # Troca ordem das colunas
    colunas = df.columns.tolist() 
    colunas[1], colunas[2] = colunas[2], colunas[1]
    df = df[colunas]
    colunas = df.columns.tolist()
    colunas[5], colunas[6] = colunas[6], colunas[5]
    df = df[colunas]
    return df

# Função para processar as Vendas BR
def processar_vendas_br(filename):
    print(f'2. Processando Vendas BR.')
    df = pd.read_excel(filename)
    df = df.iloc[:, [0, 1, 4, 8, 9, 10]]
    df.drop(index=[0, 1, 2, 3], inplace=True)
    df.columns = df.iloc[0]
    df = df.rename(columns={df.columns[2]: "Produto"})
    df = df.drop(4)
    df = df.dropna(subset=['Produto'])
    df['Vendido'] = df['Vendido'].astype(int)
    df['Estoque'] = df['Estoque'].astype(int)
    df['Recomendação BR'] = df['Vendido'] * fator - df['Estoque']
    df['Recomendação BR'] = df['Recomendação BR'].apply(math.ceil)
    df = df.rename(columns={df.columns[3]: "Vendas BR"})
    return df

# Função para processar as Vendas PY
def processar_vendas_py(filename):
    print(f'3. Processando Vendas PY.')
    df = pd.read_excel(filename)
    colunas_para_remover = [0, 2, 3, 5, 6, 7, 8, 9, 13]
    df = df.drop(df.columns[colunas_para_remover], axis=1)
    df = df.dropna(how='all')
    df.columns = df.iloc[0]
    df = df.drop(2)
    df['CANT. VENDIDA'] = df['CANT. VENDIDA'].astype(int)
    df['STOCK'] = df['STOCK'].astype(int)
    df['Recomendação PY'] = df['CANT. VENDIDA'] * fator - df['STOCK']
    df['Recomendação PY'] = df['Recomendação PY'].apply(math.ceil)
    df = df.sort_values(by=['MARCA', 'DESCRIPCION'])
    df = df.rename(columns={df.columns[0]: "Cod PY", df.columns[1]: "Produto", df.columns[2]: "Marca", df.columns[3]: "Vendas PY", df.columns[4]: "Estoque PY"})
    return df

# Função para juntar as tabelas
def juntar_tabelas(produtos, vendas_br, vendas_py):
    print(f'4. Juntando Tabelas.')
    vendas_br_selecionado = vendas_br[['Cod BR', 'Vendas BR', 'Recomendação BR']]
    vendas_py_selecionado = vendas_py[['Cod PY', 'Vendas PY', 'Estoque PY', 'Recomendação PY']]
    produtos = pd.merge(produtos, vendas_br_selecionado, on='Cod BR', how='left')
    produtos = pd.merge(produtos, vendas_py_selecionado, on='Cod PY', how='left')
    produtos = produtos.fillna({'Vendas BR': 0, 'Vendas PY': 0, 'Estoque BR': 0, 'Estoque PY': 0, 'Recomendação BR': 0, 'Recomendação PY': 0})
    return produtos

# Função para calcular se tem para o PY
def calcular_tem_para_py(produtos):
    print(f'5. Calculando se tem para o PY.')
    produtos['Tem p/ PY?'] = np.where((produtos['Recomendação PY'] > 0) & (produtos['Recomendação BR'] <= 0) & (produtos['Estoque BR'] > 0), "Sim", "Não")
    return produtos

# Função para calcular recomendações de compras
def calcular_recomendacoes_compras(produtos):
    print(f'6. Calculando Recomendações de Compras.')
    produtos['Recomendação BR'] = produtos['Vendas BR'] * fator - produtos['Estoque BR']
    # Quanto comprar? 
    # Se Filial não precisa e o Matriz sim: compra.
    # Do contrário, Matriz manda o que o Filial precisar e compra pra repor, ou compra nada.
    produtos['Quanto comprar?'] = np.where(
        (produtos['Recomendação PY'] < 0) & (produtos['Recomendação BR'] > 0) 
        , produtos['Recomendação BR'] 
        , produtos['Recomendação PY'] + produtos['Recomendação BR'])
    # Arredonda pra cima
    produtos['Vendas BR'] = produtos['Vendas BR'].apply(math.ceil)
    produtos['Recomendação BR'] = produtos['Recomendação BR'].apply(math.ceil)
    produtos['Vendas PY'] = produtos['Vendas PY'].apply(math.ceil)
    produtos['Estoque PY'] = produtos['Estoque PY'].apply(math.ceil)
    produtos['Recomendação PY'] = produtos['Recomendação PY'].apply(math.ceil)
    produtos['Quanto comprar?'] = produtos['Quanto comprar?'].apply(math.ceil)
    return produtos

# Função para calcular custo previsto
def calcular_custo_previsto(produtos):
    print(f'7. Calculando Custo Previsto.')
    produtos['Custo Previsto'] = produtos['Quanto comprar?'] * produtos['Custo']
    produtos['Comprar'] = np.where(produtos['Quanto comprar?'] > 0, produtos['Quanto comprar?'].astype(str) + " " + produtos['Unidade'] + " - " + produtos['Produto'], "")
    return produtos

def escrever_textos_recomendacao_separacao(produtos):
    print(f'8. Escrevendo os textos de Recomendação e Separação.')
    
    # Calcula a quantidade para o PY
    quantidade_py = np.where(
        (produtos['Recomendação BR'] < 0),
        np.minimum(produtos['Recomendação PY'], -produtos['Recomendação BR']),
        produtos['Recomendação PY']
    )
    
    # Converte todos os valores para strings
    quantidade_py_str = quantidade_py.astype(str)
    unidade_str = produtos['Unidade'].astype(str)
    produto_str = produtos['Produto'].astype(str)
    
    # Cria a coluna 'Separar p/ PY' usando uma abordagem segura
    produtos['Separar p/ PY'] = [
        f"{qtd} {uni} - {prod}" if (rec_py > 0 and tem_py == 'Sim') else ""
        for qtd, uni, prod, rec_py, tem_py in zip(
            quantidade_py_str, unidade_str, produto_str,
            produtos['Recomendação PY'], produtos['Tem p/ PY?']
        )
    ]
    
    # Ordena o DataFrame
    produtos = produtos.sort_values(by=['Marca', 'Produto'])
    return produtos  

def salvar_arquivo(produtos):
    # Cria uma janela root do tkinter (não será exibida)
    root = tk.Tk()
    root.withdraw()  # Esconde a janela principal do tkinter

    # Abre a janela de diálogo para salvar o arquivo
    caminho_arquivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx",  # Extensão padrão do arquivo
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],  # Tipos de arquivo permitidos
        title="Salvar como",  # Título da janela
        initialfile="Relatorio Geral de Compras.xlsx"  # Nome inicial do arquivo
    )

    # Se o usuário cancelar a janela de diálogo, retorne False
    if not caminho_arquivo:
        print("Cancelado pelo usuário.")
        return False  # Indica que o salvamento foi cancelado

    # Cria o arquivo Excel
    wb = Workbook()
    ws = wb.active

    # Define o nome da aba como a data atual no formato "DD-MM-YYYY"
    data_atual = datetime.now().strftime("%d-%m-%Y")
    ws.title = data_atual  # Define o nome da aba

    # Adiciona os dados do DataFrame à planilha
    for row in dataframe_to_rows(produtos, index=False, header=True):
        ws.append(row)

    # Formata a tabela
    tab = Table(displayName="TabelaFormatada", ref=f"A1:{chr(64 + len(produtos.columns))}{len(produtos) + 1}")
    style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Oculta as linhas de grade
    ws.sheet_view.showGridLines = False

    # Formata as células de custo como moeda
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        for cell in col:
            if cell.value in ["Custo", "Custo Previsto"]:
                col_index = cell.column
                for row in ws.iter_rows(min_row=2, max_row=len(produtos) + 1, min_col=col_index, max_col=col_index):
                    for cell in row:
                        cell.number_format = '_("R$"* #,##0.00_);_("R$"* -#,##0.00;_("R$"* "-"??_);_(@_)'
                        if isinstance(cell.value, (int, float)):
                            cell.value = float(cell.value)

    # Ajusta a largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 2) * 1.2

    # Salva o arquivo no caminho escolhido pelo usuário
    wb.save(caminho_arquivo)
    print(f"Arquivo salvo com sucesso em: {caminho_arquivo}")
    return True  # Indica que o arquivo foi salvo com sucesso

# Função para centralizar a janela
def centralizar_janela(janela, largura, altura):
    # Obtém a largura e altura da tela
    largura_tela = janela.winfo_screenwidth()
    altura_tela = janela.winfo_screenheight()

    # Calcula a posição x e y para centralizar a janela
    x = (largura_tela // 2) - (largura // 2)
    y = (altura_tela // 2) - (altura // 2)

    # Define a geometria da janela
    janela.geometry(f"{largura}x{altura}+{x}+{y}")

# Função para exibir as instruções
def mostrar_instrucoes():
    instrucoes = """
    1. Selecione os arquivos necessários

    Catálogo de Produtos

    Vendas Matriz: Relatório de Compras

    Vendas Filial: Relatório do Informes
    
    2. Clique em "Processar e Gerar Relatório".
    
    O relatório será salvo no local escolhido.

    Observações:
    - Certifique-se de que os arquivos estão no formato correto (.xls ou .xlsx).
    - Se ocorrer algum erro, uma mensagem será exibida com detalhes.
    """
    messagebox.showinfo("Instruções", instrucoes)

# Função principal com interface gráfica
def main():
    # Cria a janela principal
    root = tk.Tk()
    root.title("Gerador de Relatório de Compras | Cássio")  # Título da janela

    # Define o tamanho da janela
    largura_janela = 800
    altura_janela = 600

    # Centraliza a janela na tela
    centralizar_janela(root, largura_janela, altura_janela)

    # Cria um Canvas para servir como fundo
    canvas = tk.Canvas(root, width=largura_janela, height=altura_janela, bg="white")
    canvas.pack(fill="both", expand=True)

    # Frame para organizar os botões (sobre o Canvas)
    frame = tk.Frame(canvas, bg="white", bd=0, relief="groove")
    frame.place(relx=0.5, rely=0.5, anchor="center")
    
    # Função para obter o caminho base do executável ou script
    def obter_caminho_base():
        if getattr(sys, 'frozen', False):
            # Se estiver rodando como executável, use o diretório temporário do PyInstaller
            return sys._MEIPASS
        else:
            # Se estiver rodando como script, use o diretório do script
            return os.path.dirname(os.path.abspath(__file__))

    # Carrega a imagem Cássio.png (convertida do .ico)
    try:
        caminho_base = obter_caminho_base()  # Obtém o diretório base
        caminho_imagem = os.path.join(caminho_base, "Cássio.ico")  # Caminho da imagem
        imagem = Image.open(caminho_imagem)  # Carrega a imagem

        # Redimensiona a imagem mantendo a proporção
        largura_desejada = 200
        proporcao = largura_desejada / float(imagem.size[0])
        altura_desejada = int(float(imagem.size[1]) * proporcao)
        imagem = imagem.resize((largura_desejada, altura_desejada), Image.ANTIALIAS)

        # Converte a imagem para um formato compatível com o Tkinter
        foto = ImageTk.PhotoImage(imagem)

        # Cria um Label para exibir a imagem
        label_imagem = tk.Label(frame, image=foto, bg="white")
        label_imagem.image = foto  # Mantém uma referência para evitar que a imagem seja coletada pelo garbage collector
        label_imagem.pack(pady=10)  # Adiciona a imagem acima dos botões
    except Exception as e:
        print(f"Erro ao carregar a imagem: {e}")
        # Se a imagem não for carregada, exibe uma mensagem de erro
        label_imagem = tk.Label(frame, text="Imagem não encontrada", bg="white", fg="red")
        label_imagem.pack(pady=10)

    # Variáveis para armazenar os caminhos dos arquivos
    global filename_produtos, filename_vendas_br, filename_vendas_py
    filename_produtos = None
    filename_vendas_br = None
    filename_vendas_py = None

    # Função para alterar o estilo do botão para verde
    def botao_verde(botao):
        botao.config(bg="green", activebackground="darkgreen")  # Muda a cor de fundo para verde

    # Função para selecionar o arquivo de Catálogo de Produtos
    def selecionar_catalogo():
        global filename_produtos
        filename_produtos = selecionar_arquivo("Selecione o Catálogo de Produtos (formato .xls ou .xlsx)")
        if filename_produtos:
            label_catalogo.config(text=f"Arquivo selecionado: {filename_produtos.split('/')[-1]}")
            botao_verde(btn_catalogo)  # Altera o botão para verde

    # Função para selecionar o arquivo de Vendas BR
    def selecionar_vendas_br():
        global filename_vendas_br
        filename_vendas_br = selecionar_arquivo("Selecione o arquivo de Vendas Matriz (formato .xls ou .xlsx)")
        if filename_vendas_br:
            label_vendas_br.config(text=f"Arquivo selecionado: {filename_vendas_br.split('/')[-1]}")
            botao_verde(btn_vendas_br)  # Altera o botão para verde

    # Função para selecionar o arquivo de Vendas PY
    def selecionar_vendas_py():
        global filename_vendas_py
        filename_vendas_py = selecionar_arquivo("Selecione o arquivo de Vendas Filial (formato .xls ou .xlsx)")
        if filename_vendas_py:
            label_vendas_py.config(text=f"Arquivo selecionado: {filename_vendas_py.split('/')[-1]}")
            botao_verde(btn_vendas_py)  # Altera o botão para verde

    # Função para iniciar o processamento
    def iniciar_processamento():
        if not filename_produtos or not filename_vendas_br or not filename_vendas_py:
            messagebox.showerror("Erro", "Tente seguir as intruções")
            return

        try:
            # Inicia o processamento dos dados
            start = time.time()
            produtos = processar_catalogo_produtos(filename_produtos)
            vendas_br = processar_vendas_br(filename_vendas_br)
            vendas_py = processar_vendas_py(filename_vendas_py)
            produtos = juntar_tabelas(produtos, vendas_br, vendas_py)
            produtos = calcular_tem_para_py(produtos)
            produtos = calcular_recomendacoes_compras(produtos)
            produtos = calcular_custo_previsto(produtos)
            produtos = escrever_textos_recomendacao_separacao(produtos)
            
            # Tenta salvar o arquivo
            if not salvar_arquivo(produtos):  # Se o salvamento for cancelado
                print("Processamento interrompido: cancelado pelo usuário.")
                messagebox.showinfo("Processamento interrompido", "Cancelado pelo usuário.")
                return  # Interrompe a execução

            # Se o salvamento for bem-sucedido, continua
            end = time.time()
            tempo = end - start
            print(f"Tempo de execução: {tempo:.3f} segundos\n")

            # Exibe mensagem de sucesso
            messagebox.showinfo("Prontinho", "Relatório gerado com sucesso!")

        except Exception as e:
            # Exibe mensagem de erro caso algo dê errado
            messagebox.showerror("Erro", f"Ocorreu um erro ao gerar o relatório:\n{str(e)}")
    
    
    # Estilo dos botões
    estilo_botao = {
        "bg": "#FF0000",  # Cor de fundo
        "fg": "white",     # Cor do texto
        "font": ("Calibri", 11, "bold"),
        "borderwidth": 2,
        "relief": "groove",
        "activebackground": "#45a049",  # Cor ao clicar
    }

    # Botão de instruções
    btn_instrucoes = tk.Button(frame, text="Instruções", command=mostrar_instrucoes, **estilo_botao)
    btn_instrucoes.pack(pady=10, padx=10, fill="x")

    # Botões para selecionar os arquivos
    btn_catalogo = tk.Button(frame, text="Carregar Catálogo de Produtos", command=selecionar_catalogo, **estilo_botao)
    btn_catalogo.pack(pady=5, padx=10, fill="x" )

    label_catalogo = tk.Label(frame, text="Nenhum arquivo selecionado", fg="gray", bg="white")
    label_catalogo.pack(pady=5)

    btn_vendas_br = tk.Button(frame, text="Carregar Vendas Matriz", command=selecionar_vendas_br, **estilo_botao)
    btn_vendas_br.pack(pady=5, padx=10, fill="x")

    label_vendas_br = tk.Label(frame, text="Nenhum arquivo selecionado", fg="gray", bg="white")
    label_vendas_br.pack(pady=5)

    btn_vendas_py = tk.Button(frame, text="Carregar Vendas Filial", command=selecionar_vendas_py, **estilo_botao)
    btn_vendas_py.pack(pady=5, padx=10, fill="x")

    label_vendas_py = tk.Label(frame, text="Nenhum arquivo selecionado", fg="gray", bg="white")
    label_vendas_py.pack(pady=5)

    # Botão para iniciar o processamento
    btn_processar = tk.Button(frame, text="Processar e Gerar Relatório", command=iniciar_processamento, **estilo_botao)
    btn_processar.pack(pady=20, padx=10, fill="x")

    # Nota de rodapé
    footer = tk.Label(canvas, text="Desenvolvido por Cássio Cândido Ribeiro 2025", fg="gray", bg="white")
    footer.place(relx=0.5, rely=0.95, anchor="center")

    # Inicia a interface gráfica
    root.mainloop()

    # Define o comportamento ao fechar a janela
    def on_closing():
        root.destroy()  # Fecha a janela
        sys.exit(0)   # Encerra o processo completamente
        root.protocol("WM_DELETE_WINDOW", on_closing)  # Associa o evento de fechar à função on_closing

# Executa a função principal
if __name__ == "__main__":
    main()

# FIM